import re
import openpyxl

data_file=open('./fan2004.txt')

wb=openpyxl.Workbook()
sheet=wb.active

#列名入力
#グループ1
sheet.cell(row=1,column=1).value='Number'
sheet.cell(row=1,column=2).value='Rank'
sheet.cell(row=1,column=3).value='birthday'
sheet.cell(row=1,column=4).value='Sex'
sheet.cell(row=1,column=5).value='Age'
sheet.cell(row=1,column=6).value='Tall'
sheet.cell(row=1,column=7).value='Weight'
sheet.cell(row=1,column=8).value='BloodType'
sheet.cell(row=1,column=9).value='Win_rate'
sheet.cell(row=1,column=10).value='Wins_rate'
sheet.cell(row=1,column=11).value='1st_num'
sheet.cell(row=1,column=12).value='2nd_num'
sheet.cell(row=1,column=13).value='Runs_num'
sheet.cell(row=1,column=14).value='Yusyutu_num'
sheet.cell(row=1,column=15).value='Champ_num'
sheet.cell(row=1,column=16).value='ST_mean'

#グループ2
sheet.cell(row=1,column=17).value='Course1_entry_num'
sheet.cell(row=1,column=18).value='Course1_Wins_rate'
sheet.cell(row=1,column=19).value='Course1_ST_mean'
sheet.cell(row=1,column=20).value='Course1_Start_Rank'

sheet.cell(row=1,column=21).value='Course2_entry_num'
sheet.cell(row=1,column=22).value='Course2_Wins_rate'
sheet.cell(row=1,column=23).value='Course2_ST_mean'
sheet.cell(row=1,column=24).value='Course2_Start_Rank'

sheet.cell(row=1,column=25).value='Course3_entry_num'
sheet.cell(row=1,column=26).value='Course3_Wins_rate'
sheet.cell(row=1,column=27).value='Course3_ST_mean'
sheet.cell(row=1,column=28).value='Course3_Start_Rank'

sheet.cell(row=1,column=29).value='Course4_entry_num'
sheet.cell(row=1,column=30).value='Course4_Wins_rate'
sheet.cell(row=1,column=31).value='Course4_ST_mean'
sheet.cell(row=1,column=32).value='Course4_Start_Rank'

sheet.cell(row=1,column=33).value='Course5_entry_num'
sheet.cell(row=1,column=34).value='Course5_Wins_rate'
sheet.cell(row=1,column=35).value='Course5_ST_mean'
sheet.cell(row=1,column=36).value='Course5_Start_Rank'

sheet.cell(row=1,column=37).value='Course6_entry_num'
sheet.cell(row=1,column=38).value='Course6_Wins_rate'
sheet.cell(row=1,column=39).value='Course6_ST_mean'
sheet.cell(row=1,column=40).value='Course6_Start_Rank'

#ぐる3
sheet.cell(row=1,column=41).value='bef_Rank'
sheet.cell(row=1,column=42).value='bef_bef_Rank'
sheet.cell(row=1,column=43).value='bef_bef_bef_Rank'
sheet.cell(row=1,column=44).value='bef_Capavility_index'
sheet.cell(row=1,column=45).value='Now_Capability_index'
sheet.cell(row=1,column=46).value='Year'
sheet.cell(row=1,column=47).value='Period'
sheet.cell(row=1,column=48).value='CalcurationPeriod1'
sheet.cell(row=1,column=49).value='CalcurationPeriod2'
sheet.cell(row=1,column=50).value='GraduatePeriod'

#グル4
sheet.cell(row=1,column=51).value='Course1_1st_num'
sheet.cell(row=1,column=52).value='Course1_2nd_num'
sheet.cell(row=1,column=53).value='Course1_3rd_num'
sheet.cell(row=1,column=54).value='Course1_4th_num'
sheet.cell(row=1,column=55).value='Course1_5th_num'
sheet.cell(row=1,column=56).value='Course1_6th_num'
sheet.cell(row=1,column=57).value='Course1_F_num'
sheet.cell(row=1,column=58).value='Course1_L0_num'
sheet.cell(row=1,column=59).value='Course1_L1_num'
sheet.cell(row=1,column=60).value='Course1_K0_num'
sheet.cell(row=1,column=61).value='Course1_K1_num'
sheet.cell(row=1,column=62).value='Course1_S0_num'
sheet.cell(row=1,column=63).value='Course1_S1_num'
sheet.cell(row=1,column=64).value='Course1_S2_num'


sheet.cell(row=1,column=65).value='Course2_1st_num'
sheet.cell(row=1,column=66).value='Course2_2nd_num'
sheet.cell(row=1,column=67).value='Course2_3rd_num'
sheet.cell(row=1,column=68).value='Course2_4th_num'
sheet.cell(row=1,column=69).value='Course2_5th_num'
sheet.cell(row=1,column=70).value='Course2_6th_num'
sheet.cell(row=1,column=71).value='Course2_F_num'
sheet.cell(row=1,column=72).value='Course2_L0_num'
sheet.cell(row=1,column=73).value='Course2_L1_num'
sheet.cell(row=1,column=74).value='Course2_K0_num'
sheet.cell(row=1,column=75).value='Course2_K1_num'
sheet.cell(row=1,column=76).value='Course2_S0_num'
sheet.cell(row=1,column=77).value='Course2_S1_num'
sheet.cell(row=1,column=78).value='Course2_S2_num'


sheet.cell(row=1,column=79).value='Course3_1st_num'
sheet.cell(row=1,column=80).value='Course3_2nd_num'
sheet.cell(row=1,column=81).value='Course3_3rd_num'
sheet.cell(row=1,column=82).value='Course3_4th_num'
sheet.cell(row=1,column=83).value='Course3_5th_num'
sheet.cell(row=1,column=84).value='Course3_6th_num'
sheet.cell(row=1,column=85).value='Course3_F_num'
sheet.cell(row=1,column=86).value='Course3_L0_num'
sheet.cell(row=1,column=87).value='Course3_L1_num'
sheet.cell(row=1,column=88).value='Course3_K0_num'
sheet.cell(row=1,column=89).value='Course3_K1_num'
sheet.cell(row=1,column=90).value='Course3_S0_num'
sheet.cell(row=1,column=91).value='Course3_S1_num'
sheet.cell(row=1,column=92).value='Course3_S2_num'


sheet.cell(row=1,column=93).value='Course4_1st_num'
sheet.cell(row=1,column=94).value='Course4_2nd_num'
sheet.cell(row=1,column=95).value='Course4_3rd_num'
sheet.cell(row=1,column=96).value='Course4_4th_num'
sheet.cell(row=1,column=97).value='Course4_5th_num'
sheet.cell(row=1,column=98).value='Course4_6th_num'
sheet.cell(row=1,column=99).value='Course4_F_num'
sheet.cell(row=1,column=100).value='Course4_L0_num'
sheet.cell(row=1,column=101).value='Course4_L1_num'
sheet.cell(row=1,column=102).value='Course4_K0_num'
sheet.cell(row=1,column=103).value='Course4_K1_num'
sheet.cell(row=1,column=104).value='Course4_S0_num'
sheet.cell(row=1,column=105).value='Course4_S1_num'
sheet.cell(row=1,column=106).value='Course4_S2_num'


sheet.cell(row=1,column=107).value='Course5_1st_num'
sheet.cell(row=1,column=108).value='Course5_2nd_num'
sheet.cell(row=1,column=109).value='Course5_3rd_num'
sheet.cell(row=1,column=110).value='Course5_4th_num'
sheet.cell(row=1,column=111).value='Course5_5th_num'
sheet.cell(row=1,column=112).value='Course5_6th_num'
sheet.cell(row=1,column=113).value='Course5_F_num'
sheet.cell(row=1,column=114).value='Course5_L0_num'
sheet.cell(row=1,column=115).value='Course5_L1_num'
sheet.cell(row=1,column=116).value='Course5_K0_num'
sheet.cell(row=1,column=117).value='Course5_K1_num'
sheet.cell(row=1,column=118).value='Course5_S0_num'
sheet.cell(row=1,column=119).value='Course5_S1_num'
sheet.cell(row=1,column=120).value='Course5_S2_num'


sheet.cell(row=1,column=121).value='Course6_1st_num'
sheet.cell(row=1,column=122).value='Course6_2nd_num'
sheet.cell(row=1,column=123).value='Course6_3rd_num'
sheet.cell(row=1,column=124).value='Course6_4th_num'
sheet.cell(row=1,column=125).value='Course6_5th_num'
sheet.cell(row=1,column=126).value='Course6_6th_num'
sheet.cell(row=1,column=127).value='Course6_F_num'
sheet.cell(row=1,column=128).value='Course6_L0_num'
sheet.cell(row=1,column=129).value='Course6_L1_num'
sheet.cell(row=1,column=130).value='Course6_K0_num'
sheet.cell(row=1,column=131).value='Course6_K1_num'
sheet.cell(row=1,column=132).value='Course6_S0_num'
sheet.cell(row=1,column=133).value='Course6_S1_num'
sheet.cell(row=1,column=134).value='Course6_S2_num'

#行で処理
for i in range(1,1600+1):
   data_file=open('./fan2004.txt')
   k=i-1
   print(k)
   data_content=data_file.readlines()[k]
   data_content_regex1=re.compile(r'''(
				(\d\d\d\d)
			)''',re.VERBOSE)
   number=data_content_regex1.search(data_content).group()
   
   data_content_regex2=re.compile(r'''(
				([AB][12])
				([SH]\d\d\d\d\d\d)
				(\d)
				(\d\d)
				(\d\d\d)
				(\d\d)
				(.{1,2})
				(\d\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d\d)
				
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d\d)
				(\d\d\d)
				(\d\d\d)
				
				(.{1,2})
				(.{1,2})
				(.{1,2})
				(\d\d\d\d)
				(\d\d\d\d)
				(\d\d\d\d)
				(\d)
				(\d\d\d\d\d\d\d\d)
				(\d\d\d\d\d\d\d\d)
				(\d\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				
				(\d\d)
				(\d\d)
				(\d\d)
				(\d\d)
				)''',re.VERBOSE)
   mo=data_content_regex2.search(data_content)
   
   #エクセルに入力する
   #ぐる1
   sheet.cell(row=1+i,column=1).value=int(number)
   sheet.cell(row=1+i,column=2).value=mo.group(2)
   sheet.cell(row=1+i,column=3).value=mo.group(3)
   sheet.cell(row=1+i,column=4).value=int(mo.group(4))
   sheet.cell(row=1+i,column=5).value=int(mo.group(5))
   sheet.cell(row=1+i,column=6).value=int(mo.group(6))
   sheet.cell(row=1+i,column=7).value=int(mo.group(7))
   sheet.cell(row=1+i,column=8).value=mo.group(8)
   sheet.cell(row=1+i,column=9).value=int(mo.group(9))*0.01
   sheet.cell(row=1+i,column=10).value=int(mo.group(10))*0.01
   sheet.cell(row=1+i,column=11).value=int(mo.group(11))
   sheet.cell(row=1+i,column=12).value=int(mo.group(12))
   sheet.cell(row=1+i,column=13).value=int(mo.group(13))
   sheet.cell(row=1+i,column=14).value=int(mo.group(14))
   sheet.cell(row=1+i,column=15).value=int(mo.group(15))
   sheet.cell(row=1+i,column=16).value=int(mo.group(16))*0.01

   #ぐる2
   sheet.cell(row=1+i,column=17).value=int(mo.group(17))
   sheet.cell(row=1+i,column=18).value=int(mo.group(18))*0.01
   sheet.cell(row=1+i,column=19).value=int(mo.group(19))*0.01
   sheet.cell(row=1+i,column=20).value=int(mo.group(20))*0.01

   sheet.cell(row=1+i,column=21).value=int(mo.group(21))
   sheet.cell(row=1+i,column=23).value=int(mo.group(23))*0.01
   sheet.cell(row=1+i,column=22).value=int(mo.group(22))*0.01
   sheet.cell(row=1+i,column=24).value=int(mo.group(24))*0.01
   
   sheet.cell(row=1+i,column=25).value=int(mo.group(25))
   sheet.cell(row=1+i,column=26).value=int(mo.group(26))*0.01
   sheet.cell(row=1+i,column=27).value=int(mo.group(27))*0.01
   sheet.cell(row=1+i,column=28).value=int(mo.group(28))*0.01

   sheet.cell(row=1+i,column=29).value=int(mo.group(29))
   sheet.cell(row=1+i,column=30).value=int(mo.group(30))*0.01
   sheet.cell(row=1+i,column=31).value=int(mo.group(31))*0.01
   sheet.cell(row=1+i,column=32).value=int(mo.group(32))*0.01

   sheet.cell(row=1+i,column=33).value=int(mo.group(33))
   sheet.cell(row=1+i,column=34).value=int(mo.group(34))*0.01
   sheet.cell(row=1+i,column=35).value=int(mo.group(35))*0.01
   sheet.cell(row=1+i,column=36).value=int(mo.group(36))*0.01

   sheet.cell(row=1+i,column=37).value=int(mo.group(37))
   sheet.cell(row=1+i,column=38).value=int(mo.group(38))*0.01
   sheet.cell(row=1+i,column=39).value=int(mo.group(39))*0.01
   sheet.cell(row=1+i,column=40).value=int(mo.group(40))*0.01


   #ぐる3
   sheet.cell(row=1+i,column=41).value=mo.group(41)
   sheet.cell(row=1+i,column=42).value=mo.group(42)
   sheet.cell(row=1+i,column=43).value=mo.group(43)
   sheet.cell(row=1+i,column=44).value=int(mo.group(44))*0.01
   sheet.cell(row=1+i,column=45).value=int(mo.group(45))*0.01
   sheet.cell(row=1+i,column=46).value=int(mo.group(46))
   sheet.cell(row=1+i,column=47).value=int(mo.group(47))
   sheet.cell(row=1+i,column=48).value=int(mo.group(48))
   sheet.cell(row=1+i,column=49).value=int(mo.group(49))
   sheet.cell(row=1+i,column=50).value=int(mo.group(50))

   #グル4
   for j in range(51,134+1):
      sheet.cell(row=1+i,column=j).value=int(mo.group(j))


wb.save('boat_data.xlsx')

   
   
   
   
   
   